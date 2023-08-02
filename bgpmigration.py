import pandas as pd
import ipaddress
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import argparse
import io

# read the oldneighbors.txt file
with open('oldneighbors.txt', 'r') as file:
    lines = file.readlines()


# Set up argument parsing
parser = argparse.ArgumentParser(description='Process some file names.')
parser.add_argument('newneighbors', type=str, help='the filename of new neighbors', default=None, nargs='?')
args = parser.parse_args()

# parse the lines
parsed_data = []
for line in lines:
    line = line.strip()  # remove leading/trailing whitespace
    if not line:  # ignore empty lines
        continue
    words = re.split('\s+', line)
    # extract required information
    neighbor = words[0]
    as_number = words[2]
    up_down = words[-2]
    state_pfxrcd = words[-1]

    parsed_data.append([neighbor, as_number, up_down, state_pfxrcd])

# create dataframe
df_neighbors = pd.DataFrame(parsed_data, columns=['Neighbor', 'AS', 'Up/Down', 'State/PfxRcd'])

# read the oldinterfaces.txt file
with open('oldinterfaces.txt', 'r') as file:
    lines = file.readlines()

# parse the lines and create dictionary with p2p ip as key and interface as value
interfaces = {}
for line in lines:
    line = line.strip()  # remove leading/trailing whitespace
    if not line:  # ignore empty lines
        continue
    words = re.split('\s+', line)
    interfaces[words[1]] = words[0]

# function to check if two IPs are sequential
def is_sequential(ip1, ip2):
    return abs(int(ipaddress.IPv4Address(ip1)) - int(ipaddress.IPv4Address(ip2))) == 1

# map interfaces to neighbors
def map_interface(ip):
    for interface_ip, interface_name in interfaces.items():
        if is_sequential(ip, interface_ip):
            return interface_name
    return 'N/A'

df_neighbors['Interface'] = df_neighbors['Neighbor'].apply(map_interface)

# read the oldvrfs.txt file
with open('oldvrfs.txt', 'r') as file:
    lines = file.readlines()

# parse the lines and create dictionary with neighbor ip as key and vrf as value
vrfs = {}
for line in lines:
    line = line.strip()  # remove leading/trailing whitespace
    if not line:  # ignore empty lines
        continue
    words = re.split('\s+', line)
    vrfs[words[3][:-1]] = words[-1]  # remove comma from neighbor ip

# map vrfs to neighbors
def map_vrf(ip):
    return vrfs.get(ip, 'N/A')

df_neighbors['VRF'] = df_neighbors['Neighbor'].apply(map_vrf)

# Change this part to use the filename from arguments or an empty file
if args.newneighbors is None:
    file = io.StringIO("")
else:
    file = open(args.newneighbors, 'r')

with file:
    lines = file.readlines()

# parse the lines and create dictionary with neighbor ip as key and up/down, state/pfxrcd as value
new_neighbors = {}
for line in lines:
    line = line.strip()  # remove leading/trailing whitespace
    if not line:  # ignore empty lines
        continue
    words = re.split('\s+', line)
    new_neighbors[words[0]] = words[-2:]

# map new neighbors to old neighbors
def map_new_neighbor(ip):
    return new_neighbors.get(ip, ['N/A', 'N/A'])

df_neighbors[['Migrated Up/Down', 'Migrated State/PfxRcd']] = df_neighbors['Neighbor'].apply(map_new_neighbor).tolist()

# status formula
def status_formula(old_up_down, new_up_down, old_state_pfxrcd, new_state_pfxrcd):
    try:
        diff = int(new_state_pfxrcd) - int(old_state_pfxrcd)
        return str(diff)
    except ValueError:
        if old_state_pfxrcd in ['Idle', 'Active'] and new_state_pfxrcd in ['Idle', 'Active']:
            return 'OK'
        elif new_state_pfxrcd == 'N/A':
            return 'Not Migrated'
        elif old_state_pfxrcd in ['Idle', 'Active']:
            return "Old Migration"
        else:
            return '!!REVIEW!!'

df_neighbors['STATUS'] = df_neighbors.apply(lambda row: status_formula(row['Up/Down'], row['Migrated Up/Down'], row['State/PfxRcd'], row['Migrated State/PfxRcd']), axis=1)

# Save DataFrame to Excel
timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
filename = 'output.xlsx'

# Check if file exists, create it if it doesn't
if not os.path.exists(filename):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_neighbors.to_excel(writer, index=False, sheet_name=timestamp)
else:
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        df_neighbors.to_excel(writer, index=False, sheet_name=timestamp)

# Open the workbook and select the sheet
wb = openpyxl.load_workbook(filename)
ws = wb[timestamp]
# Open the workbook and set the width of the columns
wb = openpyxl.load_workbook(filename)
ws = wb[timestamp]  # Adjust this line to get the sheet by its name

for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Set a1 to i1 to bold and apply auto filter
bold_font = Font(bold=True)
for col in range(1, 10):
    ws[get_column_letter(col)+'1'].font = bold_font
ws.auto_filter.ref = ws.dimensions

# Define fills
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
green_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Apply color fill to the last column based on the conditions
for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
    cell = row[0]
    try:
        number = abs(int(cell.value))
    except:
        number = 0
    if cell.value == 'Not Migrated' or cell.value == '!!REVIEW!!' or number >= 100:
        cell.font = bold_font
        cell.fill = red_fill
    elif cell.value != '0':
        cell.font = bold_font
        cell.fill = yellow_fill
    elif cell.value == '0':
        cell.font = bold_font
        cell.fill = green_fill

# Save the workbook
wb.save(filename)

